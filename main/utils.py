from django.http import HttpResponse
from openpyxl import Workbook


def export_to_excel(users):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=users.xlsx'
    workbook = Workbook()
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'users'
    # Define the titles for columns
    columns = ['ID', 'Фамилия', 'Имя', 'Дата рождения', 'Возраст']
    row_num = 1
    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
    # Iterate through all movies
    for user in users:
        row_num += 1

        # Define the data for each cell in the row
        row = [
            user.pk,
            user.last_name,
            user.first_name,
            user.date_of_birth.strftime("%d.%m.%Y"),
            user.get_age(),
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    workbook.save(response)
    return response
